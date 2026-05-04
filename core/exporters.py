import csv
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def export_to_csv(queryset, fields, filename):
    """Generic CSV exporter."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    
    writer = csv.writer(response)
    writer.writerow([f.replace('_', ' ').title() for f in fields])
    
    for obj in queryset:
        row = []
        for field in fields:
            val = getattr(obj, field)
            if callable(val): val = val()
            row.append(val)
        writer.writerow(row)
    
    return response

def export_to_excel(data_list, headers, filename, sheet_name="Data"):
    """Generic Excel exporter using openpyxl."""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="00E5A0", end_color="00E5A0", fill_type="solid")
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[cell.column_letter].width = 20

    # Write data
    for row_idx, row_data in enumerate(data_list, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
            
    wb.save(response)
    return response
